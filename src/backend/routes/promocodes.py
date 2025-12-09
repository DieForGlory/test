from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import List, Optional
from datetime import datetime
import json
from io import BytesIO
from openpyxl import Workbook, load_workbook
from fastapi import Query
from database import get_db, PromoCode
from schemas import PromoCodeCreate, PromoCodeUpdate, PromoCodeResponse
from auth import require_admin

router = APIRouter()


# --- CRUD ---

@router.get("/api/admin/promocodes", response_model=List[PromoCodeResponse])
def get_promocodes(db: Session = Depends(get_db), current_user=Depends(require_admin)):
    promocodes = db.query(PromoCode).order_by(PromoCode.created_at.desc()).all()
    # Конвертация JSON строк обратно в списки для Pydantic
    for p in promocodes:
        p.applicable_products = json.loads(p.applicable_products) if p.applicable_products else []
        p.applicable_collections = json.loads(p.applicable_collections) if p.applicable_collections else []
    return promocodes


@router.get("/api/promocodes/validate")
def validate_promocode(code: str = Query(...), db: Session = Depends(get_db)):
    """Check if promo code is valid and return its details"""
    promo = db.query(PromoCode).filter(PromoCode.code == code).first()

    if not promo:
        raise HTTPException(status_code=404, detail="Промокод не найден")

    if not promo.active:
        raise HTTPException(status_code=400, detail="Промокод неактивен")

    now = datetime.utcnow()
    if promo.valid_from and promo.valid_from > now:
        raise HTTPException(status_code=400, detail="Срок действия промокода еще не начался")

    if promo.valid_until and promo.valid_until < now:
        raise HTTPException(status_code=400, detail="Срок действия промокода истек")

    # Возвращаем данные, необходимые для расчета на клиенте
    return {
        "code": promo.code,
        "discount_percent": promo.discount_percent,
        "applicable_products": json.loads(promo.applicable_products) if promo.applicable_products else [],
        "applicable_collections": json.loads(promo.applicable_collections) if promo.applicable_collections else []
    }
@router.get("/api/admin/promocodes/{id}", response_model=PromoCodeResponse)
def get_promocode(id: int, db: Session = Depends(get_db), current_user=Depends(require_admin)):
    promo = db.query(PromoCode).filter(PromoCode.id == id).first()
    if not promo:
        raise HTTPException(status_code=404, detail="Promo code not found")

    promo.applicable_products = json.loads(promo.applicable_products) if promo.applicable_products else []
    promo.applicable_collections = json.loads(promo.applicable_collections) if promo.applicable_collections else []
    return promo


@router.post("/api/admin/promocodes", response_model=PromoCodeResponse)
def create_promocode(data: PromoCodeCreate, db: Session = Depends(get_db), current_user=Depends(require_admin)):
    existing = db.query(PromoCode).filter(PromoCode.code == data.code).first()
    if existing:
        raise HTTPException(status_code=400, detail="Code already exists")

    new_promo = PromoCode(
        code=data.code,
        discount_percent=data.discount_percent,
        valid_from=data.valid_from or datetime.utcnow(),
        valid_until=data.valid_until,
        applicable_products=json.dumps(data.applicable_products),
        applicable_collections=json.dumps(data.applicable_collections),
        active=data.active
    )
    db.add(new_promo)
    db.commit()
    db.refresh(new_promo)

    # Для ответа pydantic
    new_promo.applicable_products = data.applicable_products
    new_promo.applicable_collections = data.applicable_collections
    return new_promo


@router.put("/api/admin/promocodes/{id}")
def update_promocode(id: int, data: PromoCodeUpdate, db: Session = Depends(get_db),
                     current_user=Depends(require_admin)):
    promo = db.query(PromoCode).filter(PromoCode.id == id).first()
    if not promo:
        raise HTTPException(status_code=404, detail="Not found")

    promo.code = data.code
    promo.discount_percent = data.discount_percent
    promo.valid_from = data.valid_from
    promo.valid_until = data.valid_until
    promo.applicable_products = json.dumps(data.applicable_products)
    promo.applicable_collections = json.dumps(data.applicable_collections)
    promo.active = data.active

    db.commit()
    return {"message": "Updated successfully"}


@router.delete("/api/admin/promocodes/{id}")
def delete_promocode(id: int, db: Session = Depends(get_db), current_user=Depends(require_admin)):
    promo = db.query(PromoCode).filter(PromoCode.id == id).first()
    if not promo:
        raise HTTPException(status_code=404, detail="Not found")
    db.delete(promo)
    db.commit()
    return {"message": "Deleted successfully"}


# --- Excel Import/Export ---

@router.get("/api/admin/promocodes/export/excel")
def export_promocodes(db: Session = Depends(get_db), current_user=Depends(require_admin)):
    promos = db.query(PromoCode).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Promo Codes"

    headers = ["ID", "Code", "Percent", "Start Date", "End Date", "Product IDs", "Collection IDs", "Active"]
    ws.append(headers)

    for p in promos:
        products = json.loads(p.applicable_products)
        collections = json.loads(p.applicable_collections)

        ws.append([
            p.id,
            p.code,
            p.discount_percent,
            p.valid_from.strftime("%Y-%m-%d %H:%M:%S") if p.valid_from else "",
            p.valid_until.strftime("%Y-%m-%d %H:%M:%S") if p.valid_until else "",
            ",".join(products),
            ",".join(collections),
            "Yes" if p.active else "No"
        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"promocodes_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.post("/api/admin/promocodes/import/excel")
async def import_promocodes(file: UploadFile = File(...), db: Session = Depends(get_db),
                            current_user=Depends(require_admin)):
    try:
        contents = await file.read()
        wb = load_workbook(BytesIO(contents))
        ws = wb.active

        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[1]: continue  # Skip if no code

            # Parsing logic
            code = str(row[1]).strip()
            percent = float(row[2])

            # Dates
            start_date = row[3]
            if isinstance(start_date, str) and start_date:
                try:
                    start_date = datetime.strptime(start_date, "%Y-%m-%d %H:%M:%S")
                except:
                    start_date = datetime.utcnow()
            elif not isinstance(start_date, datetime):
                start_date = datetime.utcnow()

            end_date = row[4]
            if isinstance(end_date, str) and end_date:
                try:
                    end_date = datetime.strptime(end_date, "%Y-%m-%d %H:%M:%S")
                except:
                    end_date = None
            elif not isinstance(end_date, datetime):
                end_date = None

            # Lists
            prod_ids = [x.strip() for x in str(row[5]).split(',')] if row[5] else []
            col_ids = [x.strip() for x in str(row[6]).split(',')] if row[6] else []

            active = str(row[7]).lower() in ['yes', 'true', '1']

            # Update or Create
            promo = db.query(PromoCode).filter(PromoCode.code == code).first()
            if promo:
                promo.discount_percent = percent
                promo.valid_from = start_date
                promo.valid_until = end_date
                promo.applicable_products = json.dumps(prod_ids)
                promo.applicable_collections = json.dumps(col_ids)
                promo.active = active
            else:
                new_promo = PromoCode(
                    code=code,
                    discount_percent=percent,
                    valid_from=start_date,
                    valid_until=end_date,
                    applicable_products=json.dumps(prod_ids),
                    applicable_collections=json.dumps(col_ids),
                    active=active
                )
                db.add(new_promo)
            count += 1

        db.commit()
        return {"message": f"Imported {count} codes successfully"}

    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Import failed: {str(e)}")