"""
Products Export/Import routes - Excel operations
"""
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
import json
from datetime import datetime

from database import get_db, Product
from auth import require_admin

router = APIRouter()

# ИСПРАВЛЕНО: Убрали дубликаты, которые теперь есть в основных колонках (Гендер, Диаметр, Механизм и т.д.)
# Оставляем только те спецификации, которых нет в фильтрах
SPEC_FIELDS = [
    "Стекло",
    "Калибр",
    "Запас хода",
    "Застёжка",
    "Страна - производитель",
    "Толщина корпуса", # Можно добавить, если нужно
    "Вес"
]

@router.get("/api/admin/products/export")
async def export_products(
    db: Session = Depends(get_db),
    current_user = Depends(require_admin)
):
    """Export all products to Excel file"""

    products = db.query(Product).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Products"

    # Define headers - основные поля БД (английские) + оставшиеся спец. поля (русские)
    headers = [
        "id", "name", "collection", "price", "image", "images",
        "description", "features", "in_stock", "stock_quantity",
        "sku", "is_featured",
        # Фильтры (основные колонки)
        "brand", "gender", "case_diameter", "strap_material",
        "movement", "case_material", "dial_color", "water_resistance",
        # SEO & Meta
        "seo_title", "seo_description", "seo_keywords",
        "fb_title", "fb_description", "created_at", "updated_at"
    ]

    # Добавляем оставшиеся текстовые характеристики
    headers.extend(SPEC_FIELDS)

    # Стилизация заголовков
    header_fill = PatternFill(start_color="C8102E", end_color="C8102E", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Запись данных
    for row_num, product in enumerate(products, 2):
        try:
            specs = json.loads(product.specs) if product.specs else {}
        except:
            specs = {}

        # 1. Основные поля
        ws.cell(row=row_num, column=1).value = product.id
        ws.cell(row=row_num, column=2).value = product.name
        ws.cell(row=row_num, column=3).value = product.collection
        ws.cell(row=row_num, column=4).value = product.price
        ws.cell(row=row_num, column=5).value = product.image
        ws.cell(row=row_num, column=6).value = product.images
        ws.cell(row=row_num, column=7).value = product.description
        ws.cell(row=row_num, column=8).value = product.features
        ws.cell(row=row_num, column=9).value = product.in_stock
        ws.cell(row=row_num, column=10).value = product.stock_quantity
        ws.cell(row=row_num, column=11).value = product.sku
        ws.cell(row=row_num, column=12).value = product.is_featured

        # 2. Поля фильтров (из колонок БД)
        ws.cell(row=row_num, column=13).value = product.brand
        ws.cell(row=row_num, column=14).value = product.gender
        ws.cell(row=row_num, column=15).value = product.case_diameter
        ws.cell(row=row_num, column=16).value = product.strap_material
        ws.cell(row=row_num, column=17).value = product.movement
        ws.cell(row=row_num, column=18).value = product.case_material
        ws.cell(row=row_num, column=19).value = product.dial_color
        ws.cell(row=row_num, column=20).value = product.water_resistance

        # 3. SEO
        ws.cell(row=row_num, column=21).value = product.seo_title
        ws.cell(row=row_num, column=22).value = product.seo_description
        ws.cell(row=row_num, column=23).value = product.seo_keywords
        ws.cell(row=row_num, column=24).value = product.fb_title
        ws.cell(row=row_num, column=25).value = product.fb_description
        ws.cell(row=row_num, column=26).value = product.created_at.isoformat() if product.created_at else ""
        ws.cell(row=row_num, column=27).value = product.updated_at.isoformat() if product.updated_at else ""

        # 4. Спецификации (из JSON)
        # Начинаем с 28-й колонки
        for col_offset, spec_field in enumerate(SPEC_FIELDS):
            col_num = 28 + col_offset
            ws.cell(row=row_num, column=col_num).value = specs.get(spec_field, "")

    # Авто-ширина колонок
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"orient_products_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.post("/api/admin/products/import")
async def import_products(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user = Depends(require_admin)
):
    """Import products from Excel file"""
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="File must be Excel format (.xlsx or .xls)")

    try:
        contents = await file.read()
        wb = load_workbook(BytesIO(contents))
        ws = wb.active

        headers = [cell.value for cell in ws[1]]

        required_headers = ["name", "collection", "price"]
        for required in required_headers:
            if required not in headers:
                raise HTTPException(status_code=400, detail=f"Missing required column: {required}")

        created_count = 0
        updated_count = 0
        errors = []

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            try:
                row_data = dict(zip(headers, row))
                if not row_data.get("name"): continue

                # Check exist
                existing_product = None
                if row_data.get("sku"):
                    existing_product = db.query(Product).filter(Product.sku == row_data["sku"]).first()
                if not existing_product and row_data.get("id"):
                    existing_product = db.query(Product).filter(Product.id == row_data["id"]).first()

                # Build specs dict (только из оставшихся в SPEC_FIELDS)
                specs = {}
                for spec_field in SPEC_FIELDS:
                    if spec_field in row_data and row_data[spec_field]:
                        specs[spec_field] = str(row_data[spec_field])

                # Features parsing
                features_raw = row_data.get("features")
                features_list = []
                if features_raw:
                    try:
                        features_list = json.loads(features_raw)
                        if not isinstance(features_list, list): features_list = [str(features_list)]
                    except:
                        features_list = [f.strip() for f in str(features_raw).split(',') if f.strip()]

                # Case Diameter Parsing
                case_diameter = None
                if row_data.get("case_diameter"):
                    try:
                        case_diameter = float(row_data.get("case_diameter"))
                    except:
                        pass

                product_data = {
                    "name": row_data["name"],
                    "collection": row_data["collection"],
                    "price": float(row_data["price"]) if row_data.get("price") else 0,
                    "image": row_data.get("image"),
                    "images": row_data.get("images"),
                    "description": row_data.get("description"),
                    "features": json.dumps(features_list, ensure_ascii=False),
                    "specs": json.dumps(specs, ensure_ascii=False),
                    "in_stock": bool(row_data.get("in_stock", True)),
                    "stock_quantity": int(row_data.get("stock_quantity", 0)) if row_data.get("stock_quantity") else 0,
                    "sku": row_data.get("sku"),
                    "is_featured": bool(row_data.get("is_featured", False)),

                    # Основные фильтры
                    "brand": row_data.get("brand", "Orient"),
                    "gender": row_data.get("gender"),
                    "case_diameter": case_diameter,
                    "strap_material": row_data.get("strap_material"),
                    "movement": row_data.get("movement"),
                    "case_material": row_data.get("case_material"),
                    "dial_color": row_data.get("dial_color"),
                    "water_resistance": row_data.get("water_resistance"),

                    # SEO
                    "seo_title": row_data.get("seo_title"),
                    "seo_description": row_data.get("seo_description"),
                    "seo_keywords": row_data.get("seo_keywords"),
                    "fb_title": row_data.get("fb_title"),
                    "fb_description": row_data.get("fb_description"),
                }

                if existing_product:
                    for key, value in product_data.items():
                        if value is not None: setattr(existing_product, key, value)
                    existing_product.updated_at = datetime.utcnow()
                    updated_count += 1
                else:
                    product_id = row_data.get("id")
                    if not product_id:
                        product_id = row_data["name"].lower().replace(" ", "-").replace("&", "and")
                        existing_id = db.query(Product).filter(Product.id == product_id).first()
                        if existing_id:
                            import uuid
                            product_id = f"{product_id}-{str(uuid.uuid4())[:8]}"

                    new_product = Product(id=product_id, **product_data)
                    db.add(new_product)
                    created_count += 1

            except Exception as e:
                errors.append(f"Row {row_num}: {str(e)}")
                continue

        db.commit()
        return {
            "success": True,
            "created": created_count,
            "updated": updated_count,
            "errors": errors,
            "message": f"Импорт завершен: создано {created_count}, обновлено {updated_count}"
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error processing file: {str(e)}")